import time

from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
import openpyxl as xl
from selenium.webdriver.common.keys import Keys

PATH = "/home/rifat/chromedriver"

driver = webdriver.Chrome(PATH)

# go to the website
driver.get("https://webshop.shimano.com.au/")

# create a xl sheet for saving data
workbook = xl.Workbook()
sheet = workbook.active
workbook.save('shimano.xlsx')

# login with credentials
username = driver.find_element_by_id("username")
username.clear()  # clear previous input text
username.send_keys("stkilda@bikefo.com.au")

password = driver.find_element_by_id("password")
password.clear()
password.send_keys("MovingForward2021")

driver.find_element_by_name("Login").click()
driver.implicitly_wait(3)  # wait for website to load

# allLinkFile = open('all links shimano', 'r')
# allLinkArr = []
# for line in allLinkFile:
#     allLinkArr.append(line)
# print(len(allLinkArr))

# bottomBrackets = allLinkArr[0]
# 38
driver.get('https://webshop.shimano.com.au/categories/_product_type/3_PRO/Z_CLOSE_OUTS')

# find all the elements that has same class name as below extract the links from them and save in links list


xlSheetRow = 1
kk = 0


def find_links_for_products():
    elements_links_list = []
    items_div = driver.find_elements_by_xpath('//div[@class="col-12 col-sm-6 col-md-4 col-lg-3 pb-3 mb-3"]')
    for item in items_div:
        element_name = item.find_element_by_xpath('a')
        element_link = element_name.get_attribute('href')
        elements_links_list.append(element_link)
    # del elements_links_list[:16]
    # del elements_links_list[1]
    # del elements_links_list[1]
    # del elements_links_list[2]
    # del elements_links_list[2]
    return elements_links_list


def add_column(c, click=1):
    global xlSheetRow
    xl_sheet_colmn = 1
    column_arr = []
    p_code = driver.find_element_by_class_name('product-code')
    column_arr.append(p_code.text)
    print("p_code: " + p_code.text)
    ean_code = driver.find_element_by_class_name('ean-code')
    column_arr.append(ean_code.text)
    print("ean-code: " + ean_code.text)
    product_des = driver.find_element_by_id('product-desc')
    des_image = ' '
    try:
        des_image = product_des.find_element_by_tag_name('img').get_attribute('src')
        print(des_image)
        # product_des.text += des_image
    except NoSuchElementException:
        print('no description image')
    column_arr.append(product_des.text + '\n' + des_image)
    print("product_des: " + product_des.text)
    # nameDiv = driver.find_element_by_class_name('row no-gutters mb-4')
    try:
        name = driver.find_element_by_xpath('//h3[@class="text-primary font-weight-bold mb-4"]')
    except NoSuchElementException:
        name = driver.find_element_by_xpath('//h2[@class="text-primary font-weight-bold mb-4"]')
    column_arr.append(name.text)
    print(name.text)
    ex_gst = driver.find_element_by_class_name('product-price')
    column_arr.append(ex_gst.text)
    print("ex_gst: " + ex_gst.text)
    try:
        rrp_div = driver.find_element_by_xpath('//div[@class="text-primary font-size-4"]')
        rrp = rrp_div.find_element_by_xpath('span').text
    except NoSuchElementException:
        try:
            rrp_div = driver.find_element_by_xpath('//div[@class="text-primary font-size-4 pb-4"]')
            rrp = rrp_div.find_element_by_xpath('span').text
        except NoSuchElementException:
            rrp = ' '
    print("rrp: " + rrp)

    availability_div = driver.find_element_by_xpath('//div[@class="mb-4"]')
    is_available = availability_div.find_element_by_tag_name('strong')
    column_arr.append(is_available.text)
    column_arr.append(c)

    # try:
    #     pdf_div = driver.find_element_by_xpath('//div[@class="btn-group h-100"]')
    #     pdf = pdf_div.find_element_by_tag_name('a').get_attribute('href')
    # except NoSuchElementException:
    pdf = ' '
    print("pdf: " + pdf)
    column_arr.append(pdf)

    image_div = driver.find_element_by_xpath(
        '//div[@class="border p-3 p-lg-4 mb-3 product-view position-relative text-center"]')
    a_tag_image_div = image_div.find_element_by_tag_name('a')
    image_link1 = a_tag_image_div.get_attribute('href')
    column_arr.append(image_link1)

    # try:
    #     driver.find_element_by_id('exploded-view').click()
    #     image2_div = driver.find_element_by_id('ev-page')
    #     img_tag = image2_div.find_element_by_tag_name('img')
    #     image2 = img_tag.get_attribute('src')
    # except NoSuchElementException:
    exploded_image = ' '
    column_arr.append(exploded_image)

    image2 = ' '
    if click <= 2:
        try:
            side_images = driver.find_elements_by_xpath(
                '//div[@class="border p-2 my-2 slick-slide slick-active"]')
            for si in side_images:
                image2 += si.get_attribute('data-av-switch-img-small') + ', \n'
        except NoSuchElementException:
            print('no side images')
        try:
            more_side_images = driver.find_elements_by_xpath('//div[@class="border p-2 my-2 slick-slide"]')
            for msi in more_side_images:
                image2 += msi.get_attribute('data-av-switch-img-small') + ', \n'
        except NoSuchElementException:
            print('no more side images')
    else:
        try:
            sidee_images = driver.find_elements_by_xpath(
                '//div[@class="border p-2 my-2"]')
            for si in sidee_images:
                image2 += si.get_attribute('data-av-switch-img-small') + ', \n'
        except NoSuchElementException:
            print('no side images')
    column_arr.append(image2)
    image3 = ' '
    column_arr.append(image3)
    column_arr.append(rrp)
    color_and_size = ' '

    try:
        product_codes_div = driver.find_element_by_xpath('//div[@data-avenue-part="product_codes"]')
        product_codes = product_codes_div.find_elements_by_class_name('pb-4')
        if len(product_codes) > 1:
            color_and_size = product_codes.pop().text
    except NoSuchElementException:
        print('no color and size')
    column_arr.append(color_and_size)

    # print(pName.text)
    for c in column_arr:
        if type(c) == str:
            sheet.cell(xlSheetRow, xl_sheet_colmn, c)
        else:
            sheet.cell(xlSheetRow, xl_sheet_colmn, c.text)
        xl_sheet_colmn += 1
        print("ROW: " + xlSheetRow.__str__())
        print("COLUMN: " + xl_sheet_colmn.__str__())
    xlSheetRow += 1
    workbook.save('shimano.xlsx')


def find_products_and_add():
    global kk
    products_div = driver.find_elements_by_xpath(
        '//div[@class="col-12 col-sm-6 col-md-6 col-lg-4 col-xl-3 d-flex '
        'justify-content-center justify-content-md-start"]')
    products_link = []
    directory = driver.find_elements_by_xpath('//li[@class="breadcrumb-item"]')
    category = ''
    type_of_product = directory.pop()
    directory.pop()
    for d in directory:
        n = d.find_element_by_xpath('.//following::a')
        category += n.text + '> '
    category += type_of_product.text
    print(category)
    #
    for p in products_div:
        name = p.find_element_by_tag_name('a')
        p_link = name.get_attribute('href')
        products_link.append(p_link)
    # if kk == 0:
    #     del products_link[:9]
    #     kk += 1
    print(products_link)
    #

    for pl in products_link:
        driver.get(pl)
        # driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + Keys.HOME)
        # add_column(category)
        try:
            available_color_div = driver.find_element_by_xpath('//div[@class="btn-group-toggle scs_colour"]')
            available_color = available_color_div.find_elements_by_tag_name('label')
            # available_sizes = []
            clicked = 0
            k = 0
            for c in available_color:
                c.click()
                clicked += 1
                time.sleep(2)
                try:
                    # if checked < 1:
                    available_size_div = driver.find_elements_by_xpath(
                        '//div[@class="btn-group btn-group-toggle flex-wrap '
                        'scs_size"]')
                    # checked += 1
                    labels = available_size_div[k].find_elements_by_tag_name('label')
                    for lb in labels:
                        lb.click()
                        clicked += 1
                        time.sleep(2)
                        add_column(category, clicked)
                except NoSuchElementException:
                    print('no more size available')
                    add_column(category, True)
                k += 1
        except NoSuchElementException:
            print('no more color available')
            add_column(category)


def repeat_until_found(links):
    for link in links:
        driver.get(link)
        try:
            while True:
                dropdown = driver.find_element_by_xpath('//div[@class="dropdown"]')
                # dropdown.click()
                # dropdown.find_element_by_xpath('//button[@data-value="48"]').click()
                find_products_and_add()
                driver.get(link)

                try:
                    row_div = driver.find_element_by_xpath('//div[@class="d-flex justify-content-center"]')
                    next_page_div = row_div.find_element_by_xpath('//ul[@class="pagination flex-wrap"]')
                    list_of_page_no = next_page_div.find_elements_by_tag_name('li')
                    if list_of_page_no[-1].get_attribute('class') == 'page-item disabled':
                        break
                    link = list_of_page_no[-1].find_element_by_tag_name('a').get_attribute('href')
                    print(link)
                    list_of_page_no[-1].click()
                except NoSuchElementException:
                    print('no next page')
                    break
        except NoSuchElementException:
            next_layer_links = find_links_for_products()
            repeat_until_found(next_layer_links)


productLinks = find_links_for_products()

repeat_until_found(productLinks)
