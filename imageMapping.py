import urllib.request

import openpyxl
from openpyxl.drawing.image import Image
from selenium import webdriver

shimanoFile = openpyxl.load_workbook("shimanoMainsite.xlsx")
sheet = shimanoFile.active

PATH = "/home/rifat/chromedriver"
driver = webdriver.Chrome(PATH)

outerRow = 1625
column = 7


imageNum = 1

print(sheet.max_row)

for x in range(1625, 1870):
    o = 79
    a = 65
    changed = False
    cell_obj = sheet.cell(row=outerRow, column=column)
    linkCellValue = str(cell_obj.value)
    links = linkCellValue.split(",")
    print(links)
    if len(links) > 1:
        links.pop()
    print("link len: " + str(len(links)))

    count = 1
    for l in links:
        print("count: " + str(count))
        count += 1
        print(l)
        if l != 'None':
            driver.get(l)
        else:
            continue
        urllib.request.urlretrieve(l, "images/" + str(imageNum))
        img = Image("images/" + str(imageNum))
        img.height = 70
        img.width = 70

        if changed:
            row = str(chr(o) + chr(a) + str(x))
            a += 1
        else:
            row = str(chr(o) + str(x))
            o += 1
        sheet.add_image(img, row)
        print(row)

        if o > 90:
            changed = True
            o = 65
        imageNum += 1
        shimanoFile.save("shimanoMainsite.xlsx")
    outerRow += 1


# hello = "https://www.w3schools.com/python/trypython.asp?filename=demo_ref_string_split,\n" \
#         "https://docs.google.com/document/d/1qPU7V-wJeZKlQH5sI2gDsBhydrPrLRyRaBEQ-ly-UwA/edit"

# num = 3
# print(str(chr(79) + str(num)))
# print(links)
