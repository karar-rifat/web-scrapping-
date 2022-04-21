import time

from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
import openpyxl as xl

PATH = "/home/rifat/chromedriver"
driver = webdriver.Chrome(PATH)

workbook = xl.Workbook()
sheet = workbook.active
workbook.save('shimano.xlsx')
xlSheetRow = 1


def get_image():
    image_div = driver.find_element_by_xpath('//div[@class="scalable-image"]')
    image_link = image_div.find_element_by_tag_name('img').get_attribute('src')
    return image_link


def add_column(heading_category: str, c_type=' '):
    global xlSheetRow
    xl_sheet_colmn = 1
    column_arr = []

    product_section_div = driver.find_element_by_xpath('//div[@class="product-section"]')
    product_section_contents = product_section_div.find_elements_by_class_name('product-section-content')
    left_section = product_section_contents[0]
    right_section = product_section_contents[1]
    product_code = right_section.find_element_by_class_name('product-head').text
    column_arr.append(product_code)
    print(product_code)
    name = left_section.find_element_by_tag_name('h1').text
    column_arr.append(name)
    print(name)
    product_description_div = right_section.find_element_by_class_name('product-description-area')
    product_description_p = product_description_div.find_elements_by_tag_name('p')
    product_description = ''
    for pdd in product_description_p:
        product_description += pdd.text + '\n'

    try:
        product_feature = product_description_div.find_element_by_xpath('//ul[@class="product-feature-list bft"]')
        product_feature_ul = product_feature.find_elements_by_tag_name('li')
        for pfu in product_feature_ul:
            product_description += pfu.text + '\n'
    except NoSuchElementException:
        product_description = ' '
        print('no feature available')
    column_arr.append(product_description)
    print(product_description)

    pdf_link = ''
    try:
        product_manual_li = product_description_div.find_element_by_xpath(
            '//ul[@class="btn-group"]').find_elements_by_tag_name('li')
        for pml in product_manual_li:
            pdf_name = pml.find_element_by_tag_name('a').text
            plink = pml.find_element_by_tag_name('a').get_attribute('href')
            pdf_link += pdf_name + '\n' + plink + '\n'
    except NoSuchElementException:
        pdf_link = ' '
    column_arr.append(pdf_link)
    print(pdf_link)

    technologies = ''
    try:
        technology_div = product_description_div.find_element_by_id('techLogoLink')
        li_tags = technology_div.find_elements_by_tag_name('li')
        for lit in li_tags:
            t_name = lit.find_element_by_tag_name('a').find_element_by_tag_name('img').get_attribute('alt')
            technologies += t_name + '\n'
    except NoSuchElementException:
        technologies = ' '
    column_arr.append(technologies)
    print(technologies)

    product_specification = ''
    try:
        collective_wrap_accord = driver.find_element_by_xpath('//div[@class="collective-wrap accord"]')
        table_list1 = collective_wrap_accord.find_element_by_xpath('//table[@class="spec-table"]'). \
            find_element_by_tag_name('tbody').find_elements_by_tag_name('tr')
        for tr in table_list1:
            th = tr.find_element_by_tag_name('th').text
            td = tr.find_element_by_tag_name('td').text
            product_specification += th + ': ' + td + '\n'

        try:
            expand_button = collective_wrap_accord.find_element_by_xpath('//button[@class="btn-accordion"]')
            driver.execute_script("arguments[0].click();", expand_button)
            print('expanded')
            table_list2 = collective_wrap_accord.find_element_by_xpath('//div[@class="accordion-parts open"]'). \
                find_element_by_class_name('accordion-contents'). \
                find_element_by_class_name('spec-table'). \
                find_element_by_tag_name('tbody').find_elements_by_tag_name('tr')

            for tr in table_list2:
                th = tr.find_element_by_tag_name('th').text
                td = tr.find_element_by_tag_name('td').text
                product_specification += th + ': ' + td + '\n'
        except NoSuchElementException:
            print('no expand button')
    except NoSuchElementException:
        product_specification = ' '
        print('no product specification available')
    print(product_specification)
    column_arr.append(product_specification)

    image_link = ''
    try:
        side_image_list_ul = left_section.find_elements_by_xpath(
            '//ul[@class="collective-list col3 scalable-image-list js-switch-image-list-angle"]')
        for silu in side_image_list_ul:
            side_image_list = silu.find_elements_by_tag_name('li')
            for sil in side_image_list:
                i_link = sil.find_element_by_tag_name('a').find_element_by_tag_name('img').get_attribute('src')
                image_link += i_link + ',' + '\n'
    except NoSuchElementException:
        image_link += get_image()
    # try:
    #     colective_has_icon_divs = left_section.find_element_by_xpath('//div[@class="product-thumbnail-area has-icon"]') \
    #         .find_elements_by_class_name('collective-has-icon')
    #     print(';aljd;lakjdl;aksjdal;skdjas;lkdj')
    #     for chid in colective_has_icon_divs:
    #         i = chid.find_element_by_class_name('collective-list col3 js-switch-image-list-color')
    #         print('1')
    #         i2 = i.find_element_by_tag_name('li')
    #         print('2')
    #         i3 = i2.find_element_by_class_name('imageBox')
    #         print('3')
    #         i4 = i3.find_element_by_tag_name('img').get_attribute('src')
    #         print('4')
    #         # ('//p[@class="imageBox"]').find_element_by_tag_name('img').get_attribute('src')
    #
    #         image_link += i + ', ' + '\n'
    # except NoSuchElementException:
    #     print('janina')

    try:
        more_side_images_li = left_section.find_element_by_xpath(
            '//ul[@class="collective-list col4 thumbnails-inline"]') \
            .find_elements_by_tag_name('li')
        for msil in more_side_images_li:
            i_link = msil.find_element_by_tag_name('a').find_element_by_tag_name('img').get_attribute('src')
            image_link += i_link + ', ' + '\n'
    except NoSuchElementException:
        print('no further images')
    print(image_link)
    column_arr.append(image_link)

    category_div = driver.find_element_by_xpath('//p[@class="breadcrumb"]')
    categories = category_div.find_elements_by_tag_name('a')
    del categories[0]
    c1 = categories[0].text
    print(c1)
    column_arr.append(c1)
    c2 = categories[1].text
    print(c2)
    column_arr.append(c2)
    c3 = categories[2].text
    print(c3)
    column_arr.append(c3)

    column_arr.append(heading_category)
    column_arr.append(c_type)
    print(heading_category)

    for c in column_arr:
        if type(c) == str:
            sheet.cell(xlSheetRow, xl_sheet_colmn, c)
        else:
            sheet.cell(xlSheetRow, xl_sheet_colmn, c.text)
        xl_sheet_colmn += 1
        print("ROW: " + xlSheetRow.__str__())
        print("COLUMN: " + xl_sheet_colmn.__str__())
    xlSheetRow += 1
    workbook.save('shimano.xlsx')


# ************* initial stage || extracting links **************
# footerDiv = driver.find_element_by_xpath('//ul[@class="compcat-items"]')
# footerList = footerDiv.find_elements_by_tag_name('li')
# links = []
# for f in footerList:
#     link = f.find_element_by_tag_name('a').get_attribute('href')
#     links.append(link)
# for lk in links:
#     print(lk)
# print(len(links))
# print(information)
#        *****************************
listing = ['https://bike.shimano.com/en-AU/product/apparel-accessories/s-phyre.html?filterCatId=cg5SHICEyewear',
           'https://bike.shimano.com/en-AU/apparel-accessories/eyewear/shimano-eyewear.html',

           ]

for ls in listing:
    driver.get(ls)
    driver.implicitly_wait(2)
    # headingCategory = driver.find_element_by_xpath('//h1[@class="heading-01"]').text
    # headingCategory = driver.find_element_by_xpath('//h1[@class="hero-title"]').text
    headingCategory = 'Gravel'
    # productListContentDiv = driver.find_element_by_xpath('//div[@class="js-product-list-content"]')
    productListContentDiv = driver.find_element_by_xpath('//div[@class="js-product-list-content '
                                                         'product-list-margin-top"]')
    productsDivs = productListContentDiv.find_elements_by_class_name('product-area')
    productLinks = []
    for pd in productsDivs:
        cType = pd.find_element_by_tag_name('h4').text
        liTags = pd.find_element_by_tag_name('ul').find_elements_by_tag_name('li')
        for li in liTags:
            link = li.find_element_by_class_name('product-image').find_element_by_tag_name('a').get_attribute('href')
            linkList = [cType, link]
            productLinks.append(linkList)

    for pl in productLinks:
        # print(pl.values())
        # print(pl.keys())
        driver.get(pl[1])
        cType = pl[0]
        try:
            serverError = driver.find_element_by_tag_name('h1').text
            if serverError == 'Internal Server Error':
                print('INTERNAL SERVER ERROR')
                driver.get(pl[1])
        except NoSuchElementException:
            print('ok')
        driver.implicitly_wait(2)
        add_column(headingCategory, cType)
    print(len(productLinks))
